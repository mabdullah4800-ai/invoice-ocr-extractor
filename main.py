"""
main.py
-------
Runs the whole thing: point it at a folder of invoice PDFs and it
extracts every one, validates them, and writes a single tidy
spreadsheet -- folder of PDFs in, one spreadsheet out.

    python main.py                 # processes the invoices/ folder
    python main.py path/to/folder  # processes any other folder

The spreadsheet has two sheets:

    Invoice Summary -- one row per invoice: totals, and a review status
    Line Items      -- one row per line item, ready to analyse

Anything the validator flagged is highlighted and carries a
`review_notes` column explaining exactly what to check, so a person can
go straight to the invoices that need them and ignore the rest.
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extractor import extract_invoice, get_client
from validator import validate

INVOICE_DIR = Path(__file__).parent / "invoices"
OUTPUT_PATH = Path(__file__).parent / "output" / "invoices_extracted.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2F3B52")
FLAG_FILL = PatternFill("solid", fgColor="FFE3B3")  # amber: needs a human


def process(folder: Path, client) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Extract and validate every PDF in `folder` (one API call each).

    Returns (line_items, summary) as two DataFrames.
    """
    pdfs = sorted(Path(folder).glob("*.pdf"))
    if not pdfs:
        raise FileNotFoundError(f"No PDF files found in {folder}")

    line_rows, summary_rows = [], []
    for pdf in pdfs:
        print(f"  {pdf.name} ...", end=" ", flush=True)
        extraction = extract_invoice(pdf, client)
        result = validate(extraction)
        print(result.status)

        data = extraction.data
        notes = "; ".join(result.issues)

        summary_rows.append({
            "source_file": result.source_file,
            "vendor": data.get("vendor_name"),
            "invoice_number": data.get("invoice_number"),
            "invoice_date": data.get("invoice_date"),
            "currency": data.get("currency"),
            "subtotal": data.get("subtotal"),
            "tax": data.get("tax"),
            "total": data.get("total"),
            "status": result.status,
            "review_notes": notes,
        })

        for item in data.get("line_items") or []:
            line_rows.append({
                "source_file": result.source_file,
                "invoice_number": data.get("invoice_number"),
                "vendor": data.get("vendor_name"),
                "invoice_date": data.get("invoice_date"),
                "description": item.get("description"),
                "quantity": item.get("quantity"),
                "unit_price": item.get("unit_price"),
                "line_total": item.get("line_total"),
                "needs_review": "yes" if result.status != "ok" else "",
                "review_notes": notes,
            })

    return pd.DataFrame(line_rows), pd.DataFrame(summary_rows)


def _style_sheet(sheet, flag_column: str, flag_predicate) -> None:
    """Bold the header, size the columns, and highlight flagged rows."""
    headers = [cell.value for cell in sheet[1]]
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    # Column widths from the longest value in each column.
    for col_idx, header in enumerate(headers, start=1):
        longest = max(
            [len(str(header))]
            + [len(str(sheet.cell(row=r, column=col_idx).value or ""))
               for r in range(2, sheet.max_row + 1)]
        )
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(longest + 2, 55)

    # Highlight every row the validator flagged.
    flag_idx = headers.index(flag_column)
    for row in sheet.iter_rows(min_row=2):
        if flag_predicate(row[flag_idx].value):
            for cell in row:
                cell.fill = FLAG_FILL


def write_excel(line_items: pd.DataFrame, summary: pd.DataFrame, path: Path) -> None:
    """Write the two DataFrames to a styled .xlsx workbook."""
    path.parent.mkdir(exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Invoice Summary", index=False)
        line_items.to_excel(writer, sheet_name="Line Items", index=False)

        _style_sheet(writer.book["Invoice Summary"],
                     "status", lambda v: v not in (None, "", "ok"))
        _style_sheet(writer.book["Line Items"],
                     "needs_review", lambda v: v == "yes")


def main() -> None:
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else INVOICE_DIR
    print(f"Processing invoices in {folder}/\n")

    client = get_client()
    line_items, summary = process(folder, client)
    write_excel(line_items, summary, OUTPUT_PATH)

    flagged = int((summary["status"] != "ok").sum())
    print(f"\nDone. {len(summary)} invoices processed, {flagged} need review.")
    print(f"Spreadsheet written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
